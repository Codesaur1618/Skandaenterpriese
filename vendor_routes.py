from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, current_app
from flask_login import login_required, current_user
from models import Vendor, Tenant
from forms import VendorForm
from extensions import db
from audit import log_action
from sqlalchemy import or_
from auth_routes import permission_required
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import json
import os
import csv
from pathlib import Path
from decimal import Decimal
from io import StringIO

vendor_bp = Blueprint('vendor', __name__)


@vendor_bp.route('/')
@login_required
@permission_required('view_vendors')
def list():
    tenant = Tenant.query.filter_by(code='skanda').first()
    if not tenant:
        flash('Tenant not found.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Get filter parameters
    search = request.args.get('search', '').strip()
    type_filter = request.args.get('type', '')
    credit_limit_min = request.args.get('credit_limit_min', type=float)
    credit_limit_max = request.args.get('credit_limit_max', type=float)
    
    query = Vendor.query.filter_by(tenant_id=tenant.id)
    
    # Apply filters
    if search:
        query = query.filter(
            or_(
                Vendor.name.ilike(f'%{search}%'),
                Vendor.email.ilike(f'%{search}%'),
                Vendor.contact_phone.ilike(f'%{search}%')
            )
        )
    
    if type_filter:
        query = query.filter_by(type=type_filter)
    
    if credit_limit_min is not None:
        query = query.filter(Vendor.credit_limit >= credit_limit_min)
    
    if credit_limit_max is not None:
        query = query.filter(Vendor.credit_limit <= credit_limit_max)
    
    vendors = query.order_by(Vendor.name).all()
    
    # Prepare filter data for template
    filters = [
        {
            'name': 'search',
            'label': 'Search',
            'type': 'search',
            'placeholder': 'Search by name, email, or phone...',
            'value': search,
            'icon': 'bi-search',
            'col_size': 4
        },
        {
            'name': 'type',
            'label': 'Type',
            'type': 'select',
            'value': type_filter,
            'options': [
                {'value': 'SUPPLIER', 'label': 'Supplier'},
                {'value': 'CUSTOMER', 'label': 'Customer'},
                {'value': 'BOTH', 'label': 'Both'}
            ],
            'icon': 'bi-tag',
            'col_size': 2
        },
        {
            'name': 'credit_limit',
            'label': 'Credit Limit Range',
            'type': 'number-range',
            'value_min': credit_limit_min,
            'value_max': credit_limit_max,
            'icon': 'bi-currency-rupee',
            'col_size': 3
        }
    ]
    
    # Active filters for display
    active_filters = {}
    if search:
        active_filters['Search'] = search
    if type_filter:
        active_filters['Type'] = type_filter
    if credit_limit_min is not None or credit_limit_max is not None:
        active_filters['Credit Limit'] = f"₹{credit_limit_min or 0} - ₹{credit_limit_max or '∞'}"
    
    return render_template('vendors/list.html', vendors=vendors, type_filter=type_filter,
                         filters=filters, active_filters=active_filters)


@vendor_bp.route('/new', methods=['GET', 'POST'])
@login_required
@permission_required('create_vendor')
def create():
    tenant = Tenant.query.filter_by(code='skanda').first()
    if not tenant:
        flash('Tenant not found.', 'danger')
        return redirect(url_for('vendor.list'))
    
    form = VendorForm()
    if form.validate_on_submit():
        vendor = Vendor(
            tenant_id=tenant.id,
            name=form.name.data,
            type=form.type.data,
            contact_phone=form.contact_phone.data,
            email=form.email.data,
            address=form.address.data,
            gst_number=form.gst_number.data,
            credit_limit=form.credit_limit.data or 0.00
        )
        db.session.add(vendor)
        db.session.commit()
        log_action(current_user, 'CREATE_VENDOR', 'VENDOR', vendor.id)
        flash('Vendor created successfully.', 'success')
        return redirect(url_for('vendor.list'))
    
    return render_template('vendors/form.html', form=form, title='New Vendor')


@vendor_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('edit_vendor')
def edit(id):
    vendor = Vendor.query.get_or_404(id)
    form = VendorForm(obj=vendor)
    
    if form.validate_on_submit():
        vendor.name = form.name.data
        vendor.type = form.type.data
        vendor.contact_phone = form.contact_phone.data
        vendor.email = form.email.data
        vendor.address = form.address.data
        vendor.gst_number = form.gst_number.data
        vendor.credit_limit = form.credit_limit.data or 0.00
        db.session.commit()
        log_action(current_user, 'UPDATE_VENDOR', 'VENDOR', vendor.id)
        flash('Vendor updated successfully.', 'success')
        return redirect(url_for('vendor.list'))
    
    return render_template('vendors/form.html', form=form, vendor=vendor, title='Edit Vendor')


@vendor_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('delete_vendor')
def delete(id):
    from models import Bill, ProxyBill, CreditEntry
    
    vendor = Vendor.query.get_or_404(id)
    
    # Check if vendor has associated bills
    bill_count = Bill.query.filter_by(vendor_id=vendor.id).count()
    proxy_bill_count = ProxyBill.query.filter_by(vendor_id=vendor.id).count()
    credit_count = CreditEntry.query.filter_by(vendor_id=vendor.id).count()
    
    if bill_count > 0 or proxy_bill_count > 0 or credit_count > 0:
        error_msg = f'Cannot delete vendor "{vendor.name}" because it has '
        parts = []
        if bill_count > 0:
            parts.append(f'{bill_count} bill(s)')
        if proxy_bill_count > 0:
            parts.append(f'{proxy_bill_count} proxy bill(s)')
        if credit_count > 0:
            parts.append(f'{credit_count} credit entr{"y" if credit_count == 1 else "ies"}')
        error_msg += ', '.join(parts) + ' associated with it.'
        flash(error_msg, 'danger')
        return redirect(url_for('vendor.list'))
    
    # Safe to delete - no associated records
    db.session.delete(vendor)
    db.session.commit()
    log_action(current_user, 'DELETE_VENDOR', 'VENDOR', vendor.id)
    flash('Vendor deleted successfully.', 'success')
    return redirect(url_for('vendor.list'))


@vendor_bp.route('/upload-excel', methods=['POST'])
@login_required
@permission_required('create_vendor')
def upload_excel():
    """Handle Excel or CSV file upload and import vendors"""
    tenant = Tenant.query.filter_by(code='skanda').first()
    if not tenant:
        return jsonify({'success': False, 'error': 'Tenant not found.'}), 400
    
    try:
        # Check if file is provided
        if 'excel_file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['excel_file']
        if not file or not file.filename:
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        # Check file extension
        allowed_extensions = {'xlsx', 'xls', 'csv'}
        file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        if file_ext not in allowed_extensions:
            return jsonify({'success': False, 'error': 'Invalid file type. Please upload Excel file (.xlsx, .xls) or CSV file (.csv)'}), 400
        
        # Save file temporarily
        upload_folder = Path(current_app.config.get('UPLOAD_FOLDER', Path.cwd() / 'static' / 'uploads'))
        upload_folder = upload_folder.parent / 'vendor_imports'
        upload_folder.mkdir(parents=True, exist_ok=True)
        
        filename = secure_filename(file.filename)
        timestamp = __import__('datetime').datetime.now().strftime('%Y%m%d_%H%M%S')
        filepath = upload_folder / f"import_{timestamp}_{filename}"
        
        try:
            file.save(str(filepath))
            current_app.logger.info(f"File saved to: {filepath}")
            # Verify file was saved and has content
            if not filepath.exists():
                return jsonify({'success': False, 'error': 'Failed to save uploaded file'}), 500
            if filepath.stat().st_size == 0:
                os.remove(str(filepath))
                return jsonify({'success': False, 'error': 'Uploaded file is empty'}), 400
        except Exception as e:
            current_app.logger.error(f"Error saving file: {str(e)}")
            return jsonify({'success': False, 'error': f'Failed to save file: {str(e)}'}), 500
        
        # Expected column names (matching the CSV structure exactly)
        # Note: Some columns have variations (e.g., "EMail" vs "Email", "Alternate  Mobile No." with double space)
        expected_columns = [
            'Customer Code', 'Customer Name', 'Billing Address', 'Shipping Address',
            'Pincode', 'City', 'Country', 'State', 'Status (Active/Inactive)',
            'Block Status (Yes/No)', 'Contact Person', 'Mobile No.', 'Alternate Name',
            'Alternate  Mobile No.', 'Whatsapp no.', 'EMail', 'DL20', 'DL 20 Date (From - to)',
            'DL21', 'DL 21 Date (From - to)', 'FSSAINo', 'FSSAI No 21 Date (From - to)',
            'Payment Mode', 'Credit Term (Customer/DS Type)', 'Credit Days', 'Credit Limit',
            'NoOfBillsOutstanding', 'Cust Discount', 'UID', 'RCS ID', 'Base GOI Market',
            'Market District', 'Sub-District', 'Pop Group', 'Latitude', 'Longitude',
            'Channel Type', 'Outlet Type', 'Loyalty Program', 'Service Type', 'Loyalty Tier',
            'Rev Class+T/O Class', 'GSTIN', 'PAN', 'Udhog Adhar No', 'Exemption No',
            'Trade Licence', 'Shop & Establishment Registration', 'Beat'
        ]
        
        # Also accept common variations
        column_aliases = {
            'Email': 'EMail',
            'E-Mail': 'EMail',
            'Alternate Mobile No.': 'Alternate  Mobile No.',
            'Alternate Mobile No': 'Alternate  Mobile No.',
            'WhatsApp No.': 'Whatsapp no.',
            'WhatsApp No': 'Whatsapp no.',
            'Whatsapp No.': 'Whatsapp no.',
            'FSSAI No': 'FSSAINo',
            'FSSAI No.': 'FSSAINo',
            'Customer Discount': 'Cust Discount',
            'Rev Class / T/O Class': 'Rev Class+T/O Class',
            'Rev Class+T/O Class': 'Rev Class+T/O Class',
            'Udyog Aadhaar No': 'Udhog Adhar No',
            'Udyog Aadhar No': 'Udhog Adhar No'
        }
        
        # Process file based on type
        try:
            if file_ext == 'csv':
                # Handle CSV file
                csv_data = None
                encoding_used = None
                last_error = None
                
                # First, try the simplest approach - direct file reading
                try:
                    current_app.logger.info("Trying direct CSV file reading...")
                    with open(filepath, 'r', encoding='utf-8-sig', newline='') as f:
                        csv_reader = csv.reader(f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                        # Iterate through reader instead of converting to list directly
                        csv_data = []
                        for row in csv_reader:
                            csv_data.append(row)
                        encoding_used = 'utf-8-sig'
                        current_app.logger.info(f"Direct reading with utf-8-sig succeeded, {len(csv_data)} rows")
                except UnicodeDecodeError:
                    current_app.logger.info("utf-8-sig failed, trying other encodings...")
                    csv_data = None
                except Exception as e:
                    last_error = f"Direct reading failed: {str(e)}"
                    current_app.logger.warning(last_error)
                    import traceback
                    current_app.logger.error(traceback.format_exc())
                    csv_data = None
                
                # If direct reading failed, try multiple encodings
                if csv_data is None:
                    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1', 'windows-1252']
                    delimiter = ','  # Default delimiter
                    
                    for encoding in encodings:
                        try:
                            current_app.logger.info(f"Trying encoding: {encoding}")
                            # Read file in binary mode first to check for BOM
                            with open(filepath, 'rb') as f:
                                raw_data = f.read()
                            
                            # Try to decode
                            try:
                                text_data = raw_data.decode(encoding)
                            except UnicodeDecodeError as e:
                                last_error = f"Encoding {encoding} decode failed: {str(e)}"
                                current_app.logger.warning(last_error)
                                continue
                            
                            # Try to detect delimiter from first few lines
                            lines = text_data.split('\n')[:5]
                            sample = '\n'.join(lines)
                            
                            # Count occurrences of common delimiters
                            comma_count = sample.count(',')
                            semicolon_count = sample.count(';')
                            tab_count = sample.count('\t')
                            
                            # Choose delimiter based on frequency
                            if comma_count > semicolon_count and comma_count > tab_count:
                                delimiter = ','
                            elif semicolon_count > tab_count:
                                delimiter = ';'
                            elif tab_count > 0:
                                delimiter = '\t'
                            else:
                                delimiter = ','  # Default
                            
                            # Parse CSV with proper handling
                            try:
                                csv_reader = csv.reader(
                                    StringIO(text_data),
                                    delimiter=delimiter,
                                    quotechar='"',
                                    quoting=csv.QUOTE_MINIMAL,
                                    skipinitialspace=True
                                )
                                
                                # Convert reader to list safely by iterating
                                csv_data = []
                                row_count = 0
                                for row in csv_reader:
                                    csv_data.append(row)
                                    row_count += 1
                                    # Safety limit to prevent memory issues
                                    if row_count > 100000:
                                        current_app.logger.warning("CSV file has more than 100,000 rows, processing first 100,000")
                                        break
                                
                                if csv_data and len(csv_data) > 0:
                                    encoding_used = encoding
                                    current_app.logger.info(f"Successfully read CSV with encoding {encoding}, delimiter '{delimiter}', {len(csv_data)} rows")
                                    break
                                else:
                                    csv_data = None
                            except Exception as e:
                                last_error = f"Error parsing CSV with {encoding}: {str(e)}"
                                current_app.logger.warning(last_error)
                                import traceback
                                current_app.logger.error(traceback.format_exc())
                                csv_data = None
                                continue
                            
                        except Exception as e:
                            last_error = f"Encoding {encoding} failed: {str(e)}"
                            current_app.logger.warning(last_error)
                            import traceback
                            current_app.logger.error(traceback.format_exc())
                            continue
                
                if csv_data is None or not csv_data:
                    error_msg = f'Could not read CSV file. '
                    if last_error:
                        error_msg += f'Last error: {last_error}. '
                    error_msg += 'Please ensure the file is a valid CSV with comma-separated values and proper encoding (UTF-8 recommended).'
                    try:
                        os.remove(str(filepath))
                    except:
                        pass
                    return jsonify({'success': False, 'error': error_msg}), 400
                
                if not csv_data:
                    try:
                        os.remove(str(filepath))
                    except:
                        pass
                    return jsonify({'success': False, 'error': 'CSV file is empty'}), 400
                
                # Filter out completely empty rows
                csv_data = [row for row in csv_data if any(cell and str(cell).strip() for cell in row)]
                
                if not csv_data:
                    try:
                        os.remove(str(filepath))
                    except:
                        pass
                    return jsonify({'success': False, 'error': 'CSV file contains no valid data rows'}), 400
                
                # Get header row (first row)
                headers = [str(h).strip() if h else '' for h in csv_data[0]]
                current_app.logger.info(f"CSV Headers ({len(headers)}): {headers}")
                current_app.logger.info(f"CSV First 5 headers: {headers[:5]}")
                
                # Get data rows (skip header)
                data_rows = csv_data[1:] if len(csv_data) > 1 else []
                current_app.logger.info(f"CSV Data rows: {len(data_rows)} rows found")
                
                # Log first data row for debugging
                if data_rows and len(data_rows) > 0:
                    first_row = data_rows[0]
                    current_app.logger.info(f"First data row length: {len(first_row)}")
                    current_app.logger.info(f"First data row first 5 values: {[str(v)[:50] for v in first_row[:5]]}")
                    # Check if Customer Name column exists in headers
                    if 'Customer Name' in headers:
                        name_idx = headers.index('Customer Name')
                        current_app.logger.info(f"Customer Name found at index {name_idx}, value in first row: '{first_row[name_idx] if name_idx < len(first_row) else 'OUT_OF_RANGE'}'")
                    else:
                        current_app.logger.warning("'Customer Name' not found in headers list!")
                        # Try to find it case-insensitively
                        for i, h in enumerate(headers):
                            if 'customer' in str(h).lower() and 'name' in str(h).lower():
                                current_app.logger.info(f"Found similar header at index {i}: '{h}'")
                
            else:
                # Handle Excel file
                wb = load_workbook(filepath, data_only=True)
                ws = wb.active
                
                # Get header row
                headers = [cell.value for cell in ws[1]]
                headers = [str(h).strip() if h else '' for h in headers]
                
                # Get data rows (will be processed later with iter_rows)
                data_rows = None  # Will use ws.iter_rows for Excel
            
            # Validate headers - only check for essential columns
            essential_columns = ['Customer Name']  # Only Customer Name is truly mandatory
            missing_essential = []
            
            for essential_col in essential_columns:
                found = False
                # Check in col_map (will be populated after mapping)
                # First check if it exists in headers directly
                for header in headers:
                    if essential_col.lower().strip() == str(header).lower().strip():
                        found = True
                        break
                if not found:
                    missing_essential.append(essential_col)
            
            # We'll validate column mapping after creating it, but warn if many columns are missing
            if missing_essential:
                os.remove(str(filepath))
                return jsonify({
                    'success': False,
                    'error': f'Missing essential columns: {", ".join(missing_essential)}'
                }), 400
            
            # Create column mapping (case-insensitive, flexible matching)
            col_map = {}
            
            # First pass: exact and partial matches
            for idx, header in enumerate(headers):
                if not header:  # Skip empty headers
                    continue
                    
                header_str = str(header).strip()
                header_clean = header_str.lower().replace('  ', ' ').replace('  ', ' ')  # Normalize multiple spaces
                
                # Check aliases first
                if header_str in column_aliases:
                    expected_col = column_aliases[header_str]
                    col_map[expected_col] = idx
                    current_app.logger.debug(f"Mapped '{header_str}' -> '{expected_col}' (via alias) at index {idx}")
                    continue
                
                # Try to match with expected columns
                for expected_col in expected_columns:
                    expected_clean = expected_col.lower().strip().replace('  ', ' ').replace('  ', ' ')
                    # Try exact match first
                    if expected_clean == header_clean:
                        col_map[expected_col] = idx
                        current_app.logger.debug(f"Mapped '{header_str}' -> '{expected_col}' (exact match) at index {idx}")
                        break
                    # Try partial match for variations (e.g., "Email" vs "EMail")
                    elif expected_clean.replace(' ', '').replace('/', '').replace('+', '').replace('-', '') == header_clean.replace(' ', '').replace('/', '').replace('+', '').replace('-', ''):
                        col_map[expected_col] = idx
                        current_app.logger.debug(f"Mapped '{header_str}' -> '{expected_col}' (partial match) at index {idx}")
                        break
            
            # Second pass: contains matching for critical columns (if not found in first pass)
            if 'Customer Name' not in col_map:
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    header_clean = str(header).lower().strip()
                    if 'customer' in header_clean and 'name' in header_clean:
                        col_map['Customer Name'] = idx
                        current_app.logger.info(f"Mapped '{header}' -> 'Customer Name' (contains match) at index {idx}")
                        break
            
            if 'Customer Code' not in col_map:
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    header_clean = str(header).lower().strip()
                    if 'customer' in header_clean and 'code' in header_clean:
                        col_map['Customer Code'] = idx
                        current_app.logger.info(f"Mapped '{header}' -> 'Customer Code' (contains match) at index {idx}")
                        break
            
            if 'GSTIN' not in col_map:
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    header_clean = str(header).lower().strip()
                    if 'gstin' in header_clean or ('gst' in header_clean and 'in' in header_clean):
                        col_map['GSTIN'] = idx
                        current_app.logger.info(f"Mapped '{header}' -> 'GSTIN' (contains match) at index {idx}")
                        break
            
            # Debug: Print mapping for troubleshooting
            current_app.logger.info(f"Total headers: {len(headers)}")
            current_app.logger.info(f"All headers: {headers}")  # Log ALL headers for debugging
            current_app.logger.info(f"Column mapping created: {len(col_map)} columns mapped")
            current_app.logger.info(f"Key mappings - Customer Name: {col_map.get('Customer Name', 'NOT FOUND')}, Customer Code: {col_map.get('Customer Code', 'NOT FOUND')}, GSTIN: {col_map.get('GSTIN', 'NOT FOUND')}")
            
            # Direct check: try to find Customer Name by exact string match
            if 'Customer Name' not in col_map:
                for idx, h in enumerate(headers):
                    if str(h).strip() == 'Customer Name':
                        col_map['Customer Name'] = idx
                        current_app.logger.info(f"Found 'Customer Name' by exact string match at index {idx}")
                        break
            
            # Warn if Customer Name column is not found (shouldn't happen due to earlier check, but double-check)
            if 'Customer Name' not in col_map:
                # Try to find it with different matching - be very flexible
                for idx, header in enumerate(headers):
                    header_lower = str(header).lower().strip()
                    # Try various patterns
                    if ('customer' in header_lower and 'name' in header_lower) or \
                       header_lower == 'customer name' or \
                       header_lower == 'name' or \
                       ('cust' in header_lower and 'name' in header_lower):
                        col_map['Customer Name'] = idx
                        current_app.logger.info(f"Found Customer Name at index {idx} with header: '{header}'")
                        break
                
                # Last resort: check if second column looks like a name (common pattern)
                if 'Customer Name' not in col_map and len(headers) > 1:
                    # If first column is "Customer Code" and second exists, assume it's Customer Name
                    first_header = str(headers[0]).lower().strip()
                    if 'customer' in first_header and 'code' in first_header:
                        col_map['Customer Name'] = 1
                        current_app.logger.info(f"Assumed Customer Name is at index 1 (second column)")
                
                if 'Customer Name' not in col_map:
                    os.remove(str(filepath))
                    return jsonify({
                        'success': False,
                        'error': f'Customer Name column not found in file. Available headers (first 15): {", ".join(headers[:15])}'
                    }), 400
            
            # Process rows
            results = {
                'total': 0,
                'success': 0,
                'skipped': 0,
                'errors': []
            }
            
            # Mandatory fields
            mandatory_fields = ['Customer Name']
            
            # Helper function to safely get column values
            def safe_get_value(row, col_name, default=None, row_index=None):
                col_idx = col_map.get(col_name)
                if col_idx is not None:
                    # Convert row to list if it's a tuple (Excel) or keep as list (CSV)
                    # Use type() check instead of isinstance to avoid any shadowing issues
                    if type(row).__name__ == 'list':
                        row_list = row
                    else:
                        # Convert tuple or other iterable to list
                        try:
                            row_list = [item for item in row]
                        except:
                            row_list = []
                    
                    if col_idx < len(row_list):
                        cell_value = row_list[col_idx]
                        if cell_value is not None:
                            val = str(cell_value).strip()
                            # Return value if it's not empty or common "empty" representations
                            if val and val.lower() not in ['none', 'nan', 'null', '']:
                                return val
                    else:
                        # Log if column index is out of range
                        if col_name == 'Customer Name' and row_index and row_index <= 3:
                            current_app.logger.warning(f"Row {row_index}: Column '{col_name}' index {col_idx} out of range (row has {len(row_list)} columns)")
                else:
                    # Log if column not found in mapping
                    if col_name == 'Customer Name' and row_index and row_index <= 3:
                        current_app.logger.warning(f"Row {row_index}: Column '{col_name}' not found in col_map. Available: {[k for k in col_map.keys()][:10]}")
                return default
            
            # Process rows based on file type
            if file_ext == 'csv':
                # Process CSV rows
                row_iterator = enumerate(data_rows, start=2)
            else:
                # Process Excel rows
                row_iterator = enumerate(ws.iter_rows(min_row=2, values_only=True), start=2)
            
            for row_idx, row in row_iterator:
                # Skip empty rows (handle both Excel and CSV formats)
                if file_ext == 'csv':
                    # For CSV, check if row has any non-empty values
                    if not row or not any(str(cell).strip() if cell else '' for cell in row):
                        continue
                else:
                    # For Excel, check if row has any values
                    if not any(row):
                        continue
                
                results['total'] += 1
                
                try:
                    # Extract data
                    customer_code = safe_get_value(row, 'Customer Code', row_index=row_idx)
                    customer_name = safe_get_value(row, 'Customer Name', row_index=row_idx)
                    
                    # Debug first few rows
                    if row_idx <= 3:
                        row_list = list(row) if not isinstance(row, list) else row
                        current_app.logger.info(f"Row {row_idx}: customer_name='{customer_name}', customer_code='{customer_code}', row_length={len(row_list)}")
                        if 'Customer Name' in col_map:
                            col_idx = col_map['Customer Name']
                            if col_idx < len(row_list):
                                raw_value = row_list[col_idx]
                                current_app.logger.info(f"Row {row_idx}: Customer Name column index={col_idx}, raw_value='{raw_value}', type={type(raw_value)}")
                            else:
                                current_app.logger.warning(f"Row {row_idx}: Customer Name column index {col_idx} is out of range (row has {len(row_list)} columns)")
                        else:
                            current_app.logger.error(f"Row {row_idx}: Customer Name not in col_map! Available mappings: {list(col_map.keys())[:10]}")
                    
                    # Validate mandatory fields
                    if not customer_name:
                        results['skipped'] += 1
                        error_msg = f'Row {row_idx}: Customer Name is required'
                        if 'Customer Name' in col_map:
                            col_idx = col_map['Customer Name']
                            if isinstance(row, (list, tuple)):
                                if col_idx < len(row):
                                    error_msg += f' (found: "{row[col_idx]}")'
                                else:
                                    error_msg += f' (column index {col_idx} out of range, row has {len(row)} columns)'
                        results['errors'].append(error_msg)
                        continue
                    
                    # Check for duplicates using Customer Code or GSTIN
                    gstin = safe_get_value(row, 'GSTIN', row_index=row_idx)
                    
                    # Check duplicate by customer_code
                    if customer_code:
                        existing = Vendor.query.filter_by(
                            tenant_id=tenant.id,
                            customer_code=customer_code
                        ).first()
                        if existing:
                            results['skipped'] += 1
                            results['errors'].append(f'Row {row_idx}: Duplicate Customer Code "{customer_code}" (existing vendor: {existing.name})')
                            continue
                    
                    # Check duplicate by GSTIN
                    if gstin:
                        existing = Vendor.query.filter_by(
                            tenant_id=tenant.id,
                            gst_number=gstin
                        ).first()
                        if existing:
                            results['skipped'] += 1
                            results['errors'].append(f'Row {row_idx}: Duplicate GSTIN "{gstin}" (existing vendor: {existing.name})')
                            continue
                    
                    # Extract other fields
                    def safe_get(col_name, default=''):
                        return safe_get_value(row, col_name, default, row_index=row_idx)
                    
                    # Determine vendor type (default to CUSTOMER)
                    credit_term = safe_get('Credit Term (Customer/DS Type)', '').upper()
                    vendor_type = 'CUSTOMER'  # Default
                    if 'SUPPLIER' in credit_term or 'DS' in credit_term:
                        vendor_type = 'SUPPLIER'
                    elif 'BOTH' in credit_term:
                        vendor_type = 'BOTH'
                    
                    # Get status
                    status_val = safe_get('Status (Active/Inactive)', '').upper()
                    is_active = status_val == 'ACTIVE'
                    
                    # Get credit limit
                    credit_limit_str = safe_get('Credit Limit', '0')
                    try:
                        credit_limit = Decimal(credit_limit_str.replace(',', '')) if credit_limit_str else Decimal('0.00')
                    except:
                        credit_limit = Decimal('0.00')
                    
                    # Store additional data as JSON
                    additional_data = {
                        'DL20': safe_get('DL20'),
                        'DL20_Date': safe_get('DL 20 Date (From - to)'),
                        'DL21': safe_get('DL21'),
                        'DL21_Date': safe_get('DL 21 Date (From - to)'),
                        'FSSAINo': safe_get('FSSAINo'),
                        'FSSAI_Date': safe_get('FSSAI No 21 Date (From - to)'),
                        'Payment_Mode': safe_get('Payment Mode'),
                        'Credit_Days': safe_get('Credit Days'),
                        'NoOfBillsOutstanding': safe_get('NoOfBillsOutstanding'),
                        'Cust_Discount': safe_get('Cust Discount'),
                        'UID': safe_get('UID'),
                        'RCS_ID': safe_get('RCS ID'),
                        'Base_GOI_Market': safe_get('Base GOI Market'),
                        'Market_District': safe_get('Market District'),
                        'Sub_District': safe_get('Sub-District'),
                        'Pop_Group': safe_get('Pop Group'),
                        'Latitude': safe_get('Latitude'),
                        'Longitude': safe_get('Longitude'),
                        'Channel_Type': safe_get('Channel Type'),
                        'Outlet_Type': safe_get('Outlet Type'),
                        'Loyalty_Program': safe_get('Loyalty Program'),
                        'Service_Type': safe_get('Service Type'),
                        'Loyalty_Tier': safe_get('Loyalty Tier'),
                        'Rev_Class': safe_get('Rev Class+T/O Class'),
                        'Udhog_Adhar_No': safe_get('Udhog Adhar No'),
                        'Exemption_No': safe_get('Exemption No'),
                        'Trade_Licence': safe_get('Trade Licence'),
                        'Shop_Establishment_Registration': safe_get('Shop & Establishment Registration'),
                        'Beat': safe_get('Beat')
                    }
                    
                    # Create vendor
                    vendor = Vendor(
                        tenant_id=tenant.id,
                        name=customer_name,
                        type=vendor_type,
                        customer_code=customer_code if customer_code and customer_code.lower() != 'none' and customer_code.lower() != 'nan' else None,
                        contact_phone=safe_get('Mobile No.'),
                        email=safe_get('EMail'),
                        address=safe_get('Billing Address'),
                        billing_address=safe_get('Billing Address'),
                        shipping_address=safe_get('Shipping Address'),
                        pincode=safe_get('Pincode'),
                        city=safe_get('City'),
                        country=safe_get('Country'),
                        state=safe_get('State'),
                        status=status_val if status_val else 'ACTIVE',
                        block_status=safe_get('Block Status (Yes/No)', 'NO').upper(),
                        contact_person=safe_get('Contact Person'),
                        alternate_name=safe_get('Alternate Name'),
                        alternate_mobile=safe_get('Alternate  Mobile No.'),
                        whatsapp_no=safe_get('Whatsapp no.'),
                        gst_number=gstin if gstin else None,
                        pan=safe_get('PAN') if safe_get('PAN') and safe_get('PAN').lower() != 'none' and safe_get('PAN').lower() != 'nan' else None,
                        credit_limit=credit_limit,
                        additional_data=json.dumps(additional_data) if any(additional_data.values()) else None
                    )
                    
                    db.session.add(vendor)
                    results['success'] += 1
                    
                except Exception as e:
                    results['skipped'] += 1
                    import traceback
                    error_detail = str(e)
                    # For first few errors, include more detail
                    if row_idx <= 5:
                        error_detail += f" (Traceback: {traceback.format_exc()[:200]})"
                    results['errors'].append(f'Row {row_idx}: {error_detail}')
                    current_app.logger.error(f"Error processing row {row_idx}: {traceback.format_exc()}")
                    continue
            
            # Commit all successful imports
            db.session.commit()
            
            # Log action
            log_action(current_user, 'BULK_IMPORT_VENDORS', 'VENDOR', 0)
            
            # Clean up temp file
            try:
                os.remove(str(filepath))
            except:
                pass
            
            # Prepare response message
            message = f'Import completed: {results["success"]} vendors imported, {results["skipped"]} skipped'
            if results["skipped"] > 0 and results["success"] == 0:
                # If all were skipped, show first few errors to help debug
                error_sample = results["errors"][:5] if results["errors"] else []
                if error_sample:
                    message += f'. Common errors: {"; ".join(error_sample)}'
            
            return jsonify({
                'success': True,
                'message': message,
                'results': results
            })
            
        except Exception as e:
            # Clean up temp file on error
            try:
                os.remove(str(filepath))
            except:
                pass
            return jsonify({'success': False, 'error': f'Error processing Excel file: {str(e)}'}), 500
            
    except Exception as e:
        return jsonify({'success': False, 'error': f'Upload failed: {str(e)}'}), 500

